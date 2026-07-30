"""Tests for the "export all data" archive.

The load-bearing test here is owner isolation: an export handed to the wrong person leaks a
whole portfolio at once.
"""
import io
import json
import zipfile

import pytest
from openpyxl import load_workbook

from app.services import export_service
from app.services.export_service import WORKBOOK_NAME, build_export_zip, build_workbook
from tests.conftest import OWNER_A, OWNER_B
from tests.factories import (
    make_expense_category,
    make_property,
    make_renter,
    make_supplier,
    make_transaction,
)

SHEETS = ["Properties", "Renters", "Transactions", "Suppliers", "Categories"]


def _sheets(workbook_bytes: bytes):
    return load_workbook(io.BytesIO(workbook_bytes))


def _rows(ws) -> list[dict]:
    """Sheet rows as dicts keyed by the header row."""
    values = list(ws.values)
    headers = values[0]
    return [dict(zip(headers, row)) for row in values[1:]]


@pytest.fixture
def seeded(db_session):
    """A small but complete portfolio for OWNER_A, with Hebrew text in it."""
    prop = make_property(db_session, address="רחוב הרצל 12", city="תל אביב")
    renter = make_renter(
        db_session,
        property_id=prop.id,
        first_name="שרה",
        last_name="כהן",
        extra_contacts=[{"name": "ערב", "phone": "0500000001"}],
    )
    category = make_expense_category(db_session, name="תיקונים")
    supplier = make_supplier(db_session, name="אקמה", categories=[category])
    make_transaction(db_session, property_id=prop.id, renter_id=renter.id, amount=5250.5)
    make_transaction(db_session, property_id=prop.id, categories=[category], supplier_id=supplier.id)
    return {"property": prop, "renter": renter, "category": category, "supplier": supplier}


def test_workbook_has_a_sheet_per_entity(db_session, seeded):
    wb = _sheets(build_workbook(db_session, OWNER_A))
    assert wb.sheetnames == SHEETS


def test_row_counts_match_the_seeded_data(db_session, seeded):
    wb = _sheets(build_workbook(db_session, OWNER_A))
    assert len(_rows(wb["Properties"])) == 1
    assert len(_rows(wb["Renters"])) == 1
    assert len(_rows(wb["Transactions"])) == 2
    assert len(_rows(wb["Suppliers"])) == 1


def test_hebrew_survives_the_round_trip(db_session, seeded):
    wb = _sheets(build_workbook(db_session, OWNER_A))
    assert _rows(wb["Properties"])[0]["address"] == "רחוב הרצל 12"
    renter = _rows(wb["Renters"])[0]
    assert renter["first_name"] == "שרה"
    # extra_contacts is a JSON column: dumped with ensure_ascii=False, not \u-escaped.
    assert "ערב" in renter["extra_contacts"]
    assert json.loads(renter["extra_contacts"])[0]["phone"] == "0500000001"


def test_related_names_are_resolved_next_to_their_ids(db_session, seeded):
    wb = _sheets(build_workbook(db_session, OWNER_A))
    assert _rows(wb["Renters"])[0]["property_address"] == "רחוב הרצל 12"
    expense = [r for r in _rows(wb["Transactions"]) if r["supplier_id"]][0]
    # The m2m categories column, not the legacy scalar category_id.
    assert expense["categories"] == "תיקונים"
    assert expense["supplier_name"] == "אקמה"


def test_amounts_stay_numeric(db_session, seeded):
    """openpyxl should write real numbers, so Excel can sum the column."""
    amounts = [r["amount"] for r in _rows(_sheets(build_workbook(db_session, OWNER_A))["Transactions"])]
    assert all(isinstance(a, (int, float)) for a in amounts)
    assert 5250.5 in [float(a) for a in amounts]


def test_export_does_not_leak_another_owners_data(db_session, seeded):
    """The test that matters. OWNER_B's export must contain nothing of OWNER_A's."""
    other = make_property(db_session, owner_id=OWNER_B, address="9 Secret Lane", city="Haifa")
    make_renter(db_session, owner_id=OWNER_B, property_id=other.id, first_name="Zoe", last_name="Private")
    make_transaction(db_session, owner_id=OWNER_B, property_id=other.id, amount=999.0)

    wb = _sheets(build_workbook(db_session, OWNER_B))
    blob = "\n".join(
        str(value)
        for sheet in wb.sheetnames
        for row in wb[sheet].values
        for value in row
        if value is not None
    )

    assert "9 Secret Lane" in blob  # sanity: we really did export OWNER_B
    for leaked in ("רחוב הרצל 12", "שרה", "אקמה", "תיקונים"):
        assert leaked not in blob
    assert len(_rows(wb["Transactions"])) == 1


def test_transactions_are_not_truncated_by_repository_paging(db_session, monkeypatch):
    """TransactionRepository.list() defaults to limit=100 — the export must page past it."""
    monkeypatch.setattr(export_service, "_TRANSACTION_PAGE", 5)
    prop = make_property(db_session)
    for _ in range(12):
        make_transaction(db_session, property_id=prop.id)

    wb = _sheets(build_workbook(db_session, OWNER_A))
    assert len(_rows(wb["Transactions"])) == 12


def test_zip_holds_the_workbook_and_the_owners_files(db_session, seeded, monkeypatch):
    class FakeBlob:
        def __init__(self, name, payload):
            self.name = name
            self._payload = payload

        def download_as_bytes(self):
            return self._payload

    monkeypatch.setattr(
        export_service.firebase_storage,
        "list_owner_blobs",
        lambda owner_id: [
            FakeBlob(f"{owner_id}/", b""),  # folder placeholder — skipped
            FakeBlob(f"{owner_id}/leases/lease.pdf", b"%PDF-1.4 lease"),
            FakeBlob("someone-else/private.pdf", b"nope"),  # outside the prefix — skipped
        ],
    )

    with zipfile.ZipFile(io.BytesIO(build_export_zip(db_session, OWNER_A))) as archive:
        names = archive.namelist()
        assert WORKBOOK_NAME in names
        assert names.count("files/leases/lease.pdf") == 1
        assert archive.read("files/leases/lease.pdf") == b"%PDF-1.4 lease"
        assert not any("private.pdf" in n for n in names)
        assert _sheets(archive.read(WORKBOOK_NAME)).sheetnames == SHEETS


def test_storage_failure_still_produces_a_workbook(db_session, seeded, monkeypatch):
    """Records are the part the user can't recover elsewhere — never fail the whole export."""
    def boom(_owner_id):
        raise RuntimeError("storage down")

    monkeypatch.setattr(export_service.firebase_storage, "list_owner_blobs", boom)

    with zipfile.ZipFile(io.BytesIO(build_export_zip(db_session, OWNER_A))) as archive:
        assert archive.namelist() == [WORKBOOK_NAME]
        assert len(_rows(_sheets(archive.read(WORKBOOK_NAME))["Properties"])) == 1


def test_a_single_file_failing_does_not_lose_the_others(db_session, seeded, monkeypatch):
    class Blob:
        def __init__(self, name, payload=None):
            self.name = name
            self._payload = payload

        def download_as_bytes(self):
            if self._payload is None:
                raise RuntimeError("blob gone")
            return self._payload

    monkeypatch.setattr(
        export_service.firebase_storage,
        "list_owner_blobs",
        lambda owner_id: [Blob(f"{owner_id}/broken.pdf"), Blob(f"{owner_id}/ok.pdf", b"ok")],
    )

    with zipfile.ZipFile(io.BytesIO(build_export_zip(db_session, OWNER_A))) as archive:
        assert "files/ok.pdf" in archive.namelist()
        assert "files/broken.pdf" not in archive.namelist()
