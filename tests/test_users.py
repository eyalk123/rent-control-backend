"""Full-stack tests for /users/me (account deletion cascade, data export)."""
import hashlib
import io
import zipfile

from openpyxl import load_workbook
from sqlalchemy import select

from tests.conftest import OWNER_A, OWNER_B
from tests.factories import (
    make_expense_category,
    make_property,
    make_renter,
    make_supplier,
    make_transaction,
)
from app.models.activity_log import ActivityLog
from app.models.agent import AgentConversation, AgentMessage, AgentUsageLog
from app.models.deleted_account import DeletedAccount
from app.models.device_token import DeviceToken
from app.models.document_extraction_log import DocumentExtractionLog
from app.models.expense_category import ExpenseCategory
from app.models.notification import Notification, NotificationTypeEnum
from app.models.notification_rule import NotificationRule
from app.models.property import Property
from app.models.renter import Renter
from app.models.report_export import ReportExport
from app.models.supplier import Supplier
from app.models.transaction import Transaction


def _seed_agent_data(db_session, owner_id: str) -> int:
    """A conversation + a message + a usage log for an owner. Returns the conversation id."""
    convo = AgentConversation(owner_id=owner_id, title="t")
    db_session.add(convo)
    db_session.commit()
    db_session.refresh(convo)
    db_session.add(AgentMessage(conversation_id=convo.id, role="user", content='"hi"'))
    db_session.add(
        AgentUsageLog(
            owner_id=owner_id, status="success", conversation_id=convo.id, tool_calls_count=0
        )
    )
    db_session.commit()
    return convo.id


def _seed_peripheral_data(db_session, owner_id: str) -> None:
    """The records that used to survive account deletion."""
    db_session.add_all([
        Notification(
            owner_id=owner_id,
            type=NotificationTypeEnum.OVERDUE,
            entity_id=1,
            period_key="2025-03",
        ),
        NotificationRule(
            owner_id=owner_id, event_type=NotificationTypeEnum.OVERDUE, label="rule"
        ),
        DeviceToken(owner_id=owner_id, token=f"tok-{owner_id}", platform="ios"),
        ReportExport(owner_id=owner_id, report_type="income_expense", year=2025, format="pdf"),
        DocumentExtractionLog(owner_id=owner_id, status="success"),
        ActivityLog(
            owner_id=owner_id,
            action="delete",
            entity_type="renter",
            entity_id=1,
            label="שרה כהן",
        ),
    ])
    db_session.commit()


def test_delete_account_removes_peripheral_records(client, db_session):
    """Regression: notifications, rules, device tokens, report exports, extraction logs and
    the activity log all used to outlive the account."""
    _seed_peripheral_data(db_session, OWNER_A)

    assert client.delete("/users/me").status_code == 200

    db_session.expire_all()
    for model in (Notification, NotificationRule, DeviceToken, ReportExport,
                  DocumentExtractionLog, ActivityLog):
        remaining = db_session.scalars(
            select(model).where(model.owner_id == OWNER_A)
        ).all()
        assert remaining == [], f"{model.__name__} survived account deletion"


def test_delete_account_removes_unattached_renters_and_transactions(client, db_session):
    """Regression: deletion was scoped by property id, so anything never attached to a
    property — a renter between leases, a transaction on a deleted property — survived."""
    make_renter(db_session, owner_id=OWNER_A, property_id=None)
    make_transaction(db_session, owner_id=OWNER_A, property_id=None)

    assert client.delete("/users/me").status_code == 200

    db_session.expire_all()
    assert db_session.scalars(select(Renter).where(Renter.owner_id == OWNER_A)).all() == []
    assert db_session.scalars(select(Transaction).where(Transaction.owner_id == OWNER_A)).all() == []


def test_delete_account_leaves_an_anonymous_tombstone(client, db_session):
    prop = make_property(db_session, owner_id=OWNER_A, address="רחוב הרצל 12")
    make_renter(db_session, owner_id=OWNER_A, property_id=prop.id, first_name="שרה")
    make_transaction(db_session, owner_id=OWNER_A, property_id=prop.id)

    assert client.delete("/users/me").status_code == 200

    db_session.expire_all()
    rows = db_session.scalars(select(DeletedAccount)).all()
    assert len(rows) == 1
    tombstone = rows[0]
    assert tombstone.properties_count == 1
    assert tombstone.renters_count == 1
    assert tombstone.transactions_count == 1

    # The whole point: it identifies nobody.
    assert tombstone.owner_id_hash == hashlib.sha256(OWNER_A.encode("utf-8")).hexdigest()
    assert OWNER_A not in tombstone.owner_id_hash
    values = " ".join(str(v) for v in vars(tombstone).values())
    for pii in ("רחוב הרצל 12", "שרה", OWNER_A):
        assert pii not in values


def test_delete_account_leaves_other_owners_peripheral_data(client_factory, db_session):
    _seed_peripheral_data(db_session, OWNER_B)
    _seed_peripheral_data(db_session, OWNER_A)

    assert client_factory(OWNER_A).delete("/users/me").status_code == 200

    db_session.expire_all()
    for model in (Notification, NotificationRule, DeviceToken, ReportExport,
                  DocumentExtractionLog, ActivityLog):
        assert len(db_session.scalars(
            select(model).where(model.owner_id == OWNER_B)
        ).all()) == 1


def test_export_returns_a_zip_of_the_owners_data(client, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.services.firebase_storage.list_owner_blobs", lambda owner_id: []
    )
    prop = make_property(db_session, owner_id=OWNER_A, address="1 Export Way")
    make_renter(db_session, owner_id=OWNER_A, property_id=prop.id)

    resp = client.get("/users/me/export")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"
    assert "rent-control-export-" in resp.headers["content-disposition"]

    with zipfile.ZipFile(io.BytesIO(resp.content)) as archive:
        wb = load_workbook(io.BytesIO(archive.read("rent-control-data.xlsx")))
        addresses = [row[1] for row in wb["Properties"].values][1:]
        assert addresses == ["1 Export Way"]


def test_export_is_scoped_to_the_caller(client_factory, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.services.firebase_storage.list_owner_blobs", lambda owner_id: []
    )
    make_property(db_session, owner_id=OWNER_B, address="B Only Street")
    prop_a = make_property(db_session, owner_id=OWNER_A, address="A Only Street")
    make_renter(db_session, owner_id=OWNER_A, property_id=prop_a.id)

    resp = client_factory(OWNER_A).get("/users/me/export")

    with zipfile.ZipFile(io.BytesIO(resp.content)) as archive:
        wb = load_workbook(io.BytesIO(archive.read("rent-control-data.xlsx")))
        blob = "\n".join(
            str(v) for s in wb.sheetnames for row in wb[s].values for v in row if v is not None
        )
    assert "A Only Street" in blob
    assert "B Only Street" not in blob


def test_delete_account_cascades_owner_data(client, db_session):
    prop = make_property(db_session, owner_id=OWNER_A)
    make_renter(db_session, owner_id=OWNER_A, property_id=prop.id)
    cat = make_expense_category(db_session, owner_id=OWNER_A)
    make_supplier(db_session, owner_id=OWNER_A, categories=[cat])
    make_transaction(db_session, owner_id=OWNER_A, property_id=prop.id)

    resp = client.delete("/users/me")
    assert resp.status_code == 200
    assert resp.json() == {"success": True}

    db_session.expire_all()
    assert db_session.scalars(select(Property).where(Property.owner_id == OWNER_A)).all() == []
    assert db_session.scalars(select(Renter).where(Renter.owner_id == OWNER_A)).all() == []
    assert db_session.scalars(select(Supplier).where(Supplier.owner_id == OWNER_A)).all() == []
    assert db_session.scalars(select(ExpenseCategory).where(ExpenseCategory.owner_id == OWNER_A)).all() == []
    assert db_session.scalars(select(Transaction).where(Transaction.owner_id == OWNER_A)).all() == []


def test_delete_account_leaves_other_owner_untouched(client_factory, db_session):
    prop_b = make_property(db_session, owner_id=OWNER_B)
    make_renter(db_session, owner_id=OWNER_B, property_id=prop_b.id)

    client_a = client_factory(OWNER_A)
    assert client_a.delete("/users/me").status_code == 200

    db_session.expire_all()
    assert len(db_session.scalars(select(Property).where(Property.owner_id == OWNER_B)).all()) == 1
    assert len(db_session.scalars(select(Renter).where(Renter.owner_id == OWNER_B)).all()) == 1


def test_delete_account_removes_agent_data(client, db_session):
    """Account deletion must remove the owner's chat data — portfolio PII lives in
    agent_messages verbatim."""
    convo_id = _seed_agent_data(db_session, OWNER_A)

    assert client.delete("/users/me").status_code == 200

    db_session.expire_all()
    assert db_session.scalars(
        select(AgentConversation).where(AgentConversation.owner_id == OWNER_A)
    ).all() == []
    assert db_session.scalars(
        select(AgentMessage).where(AgentMessage.conversation_id == convo_id)
    ).all() == []
    assert db_session.scalars(
        select(AgentUsageLog).where(AgentUsageLog.owner_id == OWNER_A)
    ).all() == []


def test_delete_account_leaves_other_owner_agent_data(client_factory, db_session):
    _seed_agent_data(db_session, OWNER_B)
    _seed_agent_data(db_session, OWNER_A)

    client_a = client_factory(OWNER_A)
    assert client_a.delete("/users/me").status_code == 200

    db_session.expire_all()
    assert len(db_session.scalars(
        select(AgentConversation).where(AgentConversation.owner_id == OWNER_B)
    ).all()) == 1
    assert len(db_session.scalars(
        select(AgentUsageLog).where(AgentUsageLog.owner_id == OWNER_B)
    ).all()) == 1
