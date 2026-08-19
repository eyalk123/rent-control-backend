"""Builds the owner's "export all data" archive: one .xlsx workbook + their uploaded files.

Everything here is owner-scoped — rows come from the owner-filtered repository methods, files
come from the owner's Storage prefix. Nothing in this module takes an id from the caller.
"""

import io
import json
import logging
import zipfile
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.repositories.expense_category_repository import ExpenseCategoryRepository
from app.repositories.property_repository import PropertyRepository
from app.repositories.renter_repository import RenterRepository
from app.repositories.supplier_repository import SupplierRepository
from app.repositories.transaction_repository import TransactionRepository
from app.services import firebase_storage

logger = logging.getLogger(__name__)

WORKBOOK_NAME = "rent-control-data.xlsx"

# TransactionRepository.list() paginates with limit=100 by default; an export must never
# silently truncate, so page through in batches until a short page comes back.
_TRANSACTION_PAGE = 500


def _cell(value: Any) -> Any:
    """Coerce a model attribute into something openpyxl can write."""
    if value is None or isinstance(value, (str, int, float, Decimal, bool, datetime, date)):
        # str/Enum: the enums here subclass str, so they already carry their value.
        return value.value if isinstance(value, Enum) else value
    if isinstance(value, (dict, list)):
        # extra_contacts is a JSON column. (lease_years is Text already holding JSON, so it
        # falls through as a str above.) ensure_ascii=False keeps Hebrew readable in the cell.
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _names(items) -> str:
    return ", ".join(sorted(c.name for c in items if getattr(c, "name", None)))


# Each sheet is (title, [(header, extractor)]). Explicit rather than reflected: it fixes the
# column order, keeps owner_id out of every row, and lets related names sit next to their id.
_PROPERTY_COLUMNS: list[tuple[str, Callable]] = [
    ("id", lambda p: p.id),
    ("address", lambda p: p.address),
    ("city", lambda p: p.city),
    ("zip_code", lambda p: p.zip_code),
    ("type", lambda p: p.type),
    ("floor", lambda p: p.floor),
    ("apartment", lambda p: p.apartment),
    ("block", lambda p: p.block),
    ("plot", lambda p: p.plot),
    ("number_of_rooms", lambda p: p.number_of_rooms),
    ("sq_ft", lambda p: p.sq_ft),
    ("parking_numbers", lambda p: p.parking_numbers),
    ("purchase_price", lambda p: p.purchase_price),
    ("property_owner", lambda p: p.property_owner),
    ("property_tax", lambda p: p.property_tax),
    ("house_committee", lambda p: p.house_committee),
    ("electricity_meter_number", lambda p: p.electricity_meter_number),
    ("electricity_account_number", lambda p: p.electricity_account_number),
    ("water_meter_number", lambda p: p.water_meter_number),
    ("water_account_number", lambda p: p.water_account_number),
    ("inventory_notes", lambda p: p.inventory_notes),
    ("currency_code", lambda p: p.currency_code),
    ("image_url", lambda p: p.image_url),
    ("basic_contract_url", lambda p: p.basic_contract_url),
    ("land_registry_url", lambda p: p.land_registry_url),
    ("created_at", lambda p: p.created_at),
    ("updated_at", lambda p: p.updated_at),
]

_RENTER_COLUMNS: list[tuple[str, Callable]] = [
    ("id", lambda r: r.id),
    ("first_name", lambda r: r.first_name),
    ("last_name", lambda r: r.last_name),
    ("phone", lambda r: r.phone),
    ("email", lambda r: r.email),
    ("property_id", lambda r: r.property_id),
    ("property_address", lambda r: r.property.address if r.property else None),
    ("lease_start", lambda r: r.lease_start),
    ("lease_end", lambda r: r.lease_end),
    # The binding term's end, which is what the app shows and warns on — `lease_end`
    # runs to the last option period.
    ("contract_end", lambda r: r.contract_end),
    ("terminated_on", lambda r: r.terminated_on),
    ("termination_reason", lambda r: r.termination_reason),
    ("contract_term_years", lambda r: r.contract_term_years),
    ("option_years", lambda r: r.option_years),
    ("base_rent", lambda r: r.base_rent),
    ("rent_escalation_mode", lambda r: r.rent_escalation_mode),
    ("rent_escalation_value", lambda r: r.rent_escalation_value),
    ("cpi_base_index", lambda r: r.cpi_base_index),
    ("lease_years", lambda r: r.lease_years),
    ("number_of_payments", lambda r: r.number_of_payments),
    ("payment_type", lambda r: r.payment_type),
    ("payment_day_of_month", lambda r: r.payment_day_of_month),
    ("insurance_type", lambda r: r.insurance_type),
    ("insurance_amount", lambda r: r.insurance_amount),
    ("extra_contacts", lambda r: r.extra_contacts),
    ("full_contract_url", lambda r: r.full_contract_url),
    ("id_image_url", lambda r: r.id_image_url),
    ("created_at", lambda r: r.created_at),
    ("updated_at", lambda r: r.updated_at),
]

_TRANSACTION_COLUMNS: list[tuple[str, Callable]] = [
    ("id", lambda t: t.id),
    ("type", lambda t: t.type),
    ("date_of_payment", lambda t: t.date_of_payment),
    ("month_for", lambda t: t.month_for),
    ("amount", lambda t: t.amount),
    ("currency_code", lambda t: t.currency_code),
    ("property_id", lambda t: t.property_id),
    ("property_address", lambda t: t.property_address),
    ("renter_id", lambda t: t.renter_id),
    ("renter_name", lambda t: t.renter_name),
    # The m2m is the current model; the scalar category_id is legacy.
    ("categories", lambda t: _names(t.categories)),
    ("supplier_id", lambda t: t.supplier_id),
    ("supplier_name", lambda t: t.supplier.name if t.supplier else None),
    ("payment_method", lambda t: t.payment_method),
    ("notes", lambda t: t.notes),
    ("receipt_image_url", lambda t: t.receipt_image_url),
    ("created_at", lambda t: t.created_at),
    ("updated_at", lambda t: t.updated_at),
]

_SUPPLIER_COLUMNS: list[tuple[str, Callable]] = [
    ("id", lambda s: s.id),
    ("name", lambda s: s.name),
    ("is_active", lambda s: s.is_active),
    ("phone", lambda s: s.phone),
    ("email", lambda s: s.email),
    ("bank_account", lambda s: s.bank_account),
    ("categories", lambda s: _names(s.categories)),
    ("notes", lambda s: s.notes),
    ("created_at", lambda s: s.created_at),
    ("updated_at", lambda s: s.updated_at),
]

_CATEGORY_COLUMNS: list[tuple[str, Callable]] = [
    ("id", lambda c: c.id),
    ("key", lambda c: c.key),
    ("name", lambda c: c.name),
    # Predefined categories have no owner; the owner's own additions do.
    ("built_in", lambda c: c.owner_id is None),
    ("is_active", lambda c: c.is_active),
    ("sort_order", lambda c: c.sort_order),
    ("created_at", lambda c: c.created_at),
]


def _all_transactions(db: Session, owner_id: str) -> list:
    repo = TransactionRepository(db)
    rows: list = []
    offset = 0
    while True:
        page = repo.list(owner_id=owner_id, limit=_TRANSACTION_PAGE, offset=offset)
        rows.extend(page)
        if len(page) < _TRANSACTION_PAGE:
            return rows
        offset += _TRANSACTION_PAGE


def _write_sheet(wb: Workbook, title: str, columns: list[tuple[str, Callable]], rows: list) -> None:
    ws = wb.create_sheet(title)
    ws.append([header for header, _ in columns])
    for row in rows:
        ws.append([_cell(extract(row)) for _, extract in columns])

    ws.freeze_panes = "A2"
    for index, (header, _) in enumerate(columns, start=1):
        # Rough autofit: wide enough to read, capped so a JSON blob can't blow the layout out.
        widest = max([len(header)] + [len(str(c.value)) for c in ws[get_column_letter(index)][1:] if c.value is not None])
        ws.column_dimensions[get_column_letter(index)].width = min(max(widest + 2, 10), 50)


def build_workbook(db: Session, owner_id: str) -> bytes:
    """The .xlsx half of the export: one sheet per record type, all owner-scoped."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet openpyxl creates

    _write_sheet(wb, "Properties", _PROPERTY_COLUMNS, PropertyRepository(db).get_all_by_owner(owner_id))
    _write_sheet(wb, "Renters", _RENTER_COLUMNS, RenterRepository(db).get_all(owner_id))
    _write_sheet(wb, "Transactions", _TRANSACTION_COLUMNS, _all_transactions(db, owner_id))
    _write_sheet(
        wb,
        "Suppliers",
        _SUPPLIER_COLUMNS,
        SupplierRepository(db).get_all(owner_id, include_inactive=True),
    )
    _write_sheet(
        wb,
        "Categories",
        _CATEGORY_COLUMNS,
        ExpenseCategoryRepository(db).get_all_active_ordered(owner_id),
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_export_zip(db: Session, owner_id: str) -> bytes:
    """The workbook plus every file the owner has uploaded, as a single ZIP.

    Storage problems degrade to a workbook-only archive rather than failing the export —
    the records are the part the user can't get back any other way.
    """
    workbook = build_workbook(db, owner_id)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(WORKBOOK_NAME, workbook)

        prefix = f"{owner_id}/"
        try:
            blobs = firebase_storage.list_owner_blobs(owner_id)
        except Exception as exc:
            logger.warning("Storage unavailable — exporting records only for %s: %s", owner_id, exc)
            blobs = []

        for blob in blobs:
            name = getattr(blob, "name", "") or ""
            if not name.startswith(prefix) or name.endswith("/"):
                continue  # folder placeholder, or a blob outside this owner's prefix
            try:
                archive.writestr(f"files/{name[len(prefix):]}", blob.download_as_bytes())
            except Exception as exc:
                logger.warning("Skipping file %s in export for %s: %s", name, owner_id, exc)

    return buffer.getvalue()
